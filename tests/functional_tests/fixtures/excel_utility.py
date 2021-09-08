import openpyxl
import pandas as pd
from openpyxl import load_workbook
import os


class Excel_Utility():

    def __init__(self, file_path):
        self.file_path = file_path

    def read_file(self, sheet_name='', header=0, execute='No'):
        try:
            dataframe = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header)
            # setattr(self, "read_df", dataframe)
            if execute == 'Y':
                dataframe = dataframe[dataframe['Execute'] == 'Y']
            return dataframe
        except Exception as e:
            return None

    def read_specific_data(self, start_col, start_row, end_col, end_row, sheet_name=''):
        final_datalist = []
        col_list = []
        try:
            wb_obj = openpyxl.load_workbook(self.file_path)
            sheet = wb_obj.get_sheet_by_name(sheet_name)
            df = pd.DataFrame(sheet.values)
            for row in range(start_row, end_row + 1):
                data_list = []
                for col in range(start_col, end_col + 1):
                    cell_obj = sheet.cell(row=row, column=col)
                    if row == 1:
                        col_list.append(cell_obj.value)
                    else:
                        data_list.append(cell_obj.value)
                if row != 1:
                    final_datalist.append(data_list)

            df = pd.DataFrame(final_datalist, columns=col_list)
            return df
        except Exception as e:
            print("===")
            return None

    def read_single_cell(self, row_no, col_no, sheet_name=''):
        wb_obj = openpyxl.load_workbook(self.file_path)
        sheet = wb_obj.get_sheet_by_name(sheet_name)
        df = pd.DataFrame(sheet.values)
        row_count = len(df.index)
        # df.count()
        col_count = len(df.columns)

        cell_value = sheet.cell(row=row_no, column=col_no).value
        return cell_value

    def write_into_exccel(self):
        pass

    def read_file_from_xlsx(self, sheet_name='', execute='No'):
        try:
            wb_obj = openpyxl.load_workbook(self.file_path)
            sheet = wb_obj.get_sheet_by_name(sheet_name)
            dataframe = pd.DataFrame(sheet.values)
            # setattr(self, "read_df", dataframe)
            if execute == 'Y':
                dataframe = dataframe[dataframe['Execute'] == 'Y']
            return dataframe
        except Exception as e:
            return None

# READ CSV
# df = pd.read_csv()
#
# # READ XLS USING PANDAS
# dataframe = pd.read_excel("./testdata/TDD-Automation-Info.xls", sheet_name="TC", index_col=0)
# print(dataframe['TEST_NAME'])
#
#
#
#
# # LOAD XLSX USING OPENPYXL
# wb_obj = openpyxl.load_workbook("./testdata/TDD-Automation-Info.xlsx")
# # Get workbook active sheet object
# # from the active attribute
# sheet_obj = wb_obj['TC']
# data = sheet_obj['B2'].value
# cell_obj = sheet_obj.cell(row=1, column=1)
# print(cell_obj.value)
# print("=")
# # LOAD XLSX USING OPENPYXL
#
#
#
# # Convert to DataFrame
# sheet_obj = wb_obj.active
# df = pd.DataFrame(sheet_obj.values)
#
# val1 = pd.Series('11', [10])
# val2 = pd.Series('M4', [10])
# val3 = pd.Series('abc', [10])
# val4 = pd.Series('abc xyz', [10])
# val5 = pd.Series('FAIL', [10])
#
#
# df = pd.DataFrame(val1, columns=['TC'])
# df['MODULE'] = val2
# df['TEST_NAME'] = val3
# df['TEST_DESC'] = val4
# df['STATUS'] = val5
#
# book = load_workbook("./testdata/TDD-Automation-Info.xlsx")
# if not df.empty:
#     worksheet = book.active
#
# writer = pd.ExcelWriter("./testdata/TDD-Automation-Info.xlsx", engine='openpyxl')
# writer.book = book
# # df.to_excel(writer, sheet_name="TC", startrow=writer.sheets['TC'].max_row, index=True, header=True)
# df.to_excel(writer, startrow=10, index=True, header=True)
#
#
# print("==")
# # writer.sheets = {ws.title: ws for ws in book.worksheets}
