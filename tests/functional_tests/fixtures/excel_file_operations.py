import openpyxl
import pandas as pd
from openpyxl import load_workbook

#READ CSV
# df = pd.read_csv()

#READ XLS USING PANDAS
# dataframe = pd.read_excel("./testdata/TDD-Automation-Info.xls", sheet_name="TC", index_col=0)
# print(dataframe['TEST_NAME'])
#
#
#
#
# #LOAD XLSX USING OPENPYXL
wb_obj = openpyxl.load_workbook("./testdata/TDD-Automation-Info.xlsx")
# # Get workbook active sheet object
# # from the active attribute
# sheet_obj = wb_obj['TC']
# data = sheet_obj['B2'].value
# cell_obj = sheet_obj.cell(row=1, column=1)
# print(cell_obj.value)
# print("=")
#LOAD XLSX USING OPENPYXL



# Convert to DataFrame
sheet_obj = wb_obj.active
df = pd.DataFrame(sheet_obj.values)

val1 = pd.Series('11', [10])
val2 = pd.Series('M4', [10])
val3 = pd.Series('abc', [10])
val4 = pd.Series('abc xyz', [10])
val5 = pd.Series('FAIL', [10])


df = pd.DataFrame(val1, columns=['TC'])
df['MODULE'] = val2
df['TEST_NAME'] = val3
df['TEST_DESC'] = val4
df['STATUS'] = val5

book = load_workbook("./testdata/TDD-Automation-Info.xlsx")
if not df.empty:
    worksheet = book.active

writer = pd.ExcelWriter("./testdata/TDD-Automation-Info.xlsx", engine='openpyxl')
writer.book = book
# df.to_excel(writer, sheet_name="TC", startrow=writer.sheets['TC'].max_row, index=True, header=True)
df.to_excel(writer, startrow=10, index=True, header=True)


print("==")
# writer.sheets = {ws.title: ws for ws in book.worksheets}
