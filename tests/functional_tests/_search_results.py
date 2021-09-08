import os
import openpyxl
import pandas as pd
import pytest
from tests.functional_tests.pages.home_page import Home_Page
from xlrd import book

test_data_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '.', "./fixtures/testdata/master_data.xlsx"))
dataframe = pd.read_excel(test_data_path,sheet_name="Test_Search_Results", header = 0, index_col=None)
# wb_obj = openpyxl.load_workbook(test_data_path)
# sheet = wb_obj.get_sheet_by_name("Test_Search_Results", header = 0, index_col=None)
# dataframe = pd.DataFrame(sheet.values)
dataframe = dataframe[dataframe["Execute"] == "YES"]

testParams = []
# testId = []
testName = []
for i, row in dataframe.iterrows():
    testParams.append((row['Filter'], row['Model_ID'] if row['Filter_Needed'] == 'Y' else 'No filter', i)),
    testName.append(row['Test_Name'])
    # testId.append(i)


class Test_Search_Results(object):


    def setup_class(self):
        # self.home_page = Home_Page(self.browser, self.master_data_sheet)
        # self.data_frame = self.home_page.xlsx_utility.read_file_from_xlsx(sheet_name=os.environ.get('PYTEST_CURRENT_TEST').split('::')[1], execute='Y')
        # self.data_frame.fillna("", inplace=True)
        pass

    # Decorators
    @pytest.fixture
    def clean_up(self):
        print("@@@inside fixture")


    @pytest.mark.parametrize("Filter, Model_ID, testId", testParams, ids=testName)
    @pytest.mark.usefixtures('clean_up') # 1st way of calling fixture
    def test_search_product(self, Filter, Model_ID, testId):
        try:
            print("inside test_product")
            status_series = []
            index_series = []
            status = ''
            print(Filter, Model_ID)

            if Filter != 'Filter5':
                status = 'Pass'
            else :
                status = 'Fail'
            myworkbook = openpyxl.load_workbook(test_data_path)
            worksheet = myworkbook.get_sheet_by_name('Test_Search_Results')
            worksheet.cell(row= int(testId)+2, column=8).value = status
            myworkbook.save(test_data_path)

        except Exception as e:
            print("####")





    # status_series.append('Fail') if Filter == 'Filter5' else status_series.append('Pass')
    # index_series.append(counter)



    # from openpyxl import load_workbook
    # status_column = pd.Series(status_series, index=index_series)
    # status_dataframe = pd.DataFrame(status_column, columns=['Status'])


    # with pd.ExcelWriter(test_data_path, engine='openpyxl') as writer:
    #     writer.book = load_workbook(test_data_path)
    #
    #     # worksheet = workbook.get_sheet_by_name('Test_Search_Results')
    #     # status_cell = worksheet['Status']
    #     # status_cell.value =
    #     status_dataframe.to_excel(writer, sheet_name='Test_Search_Results', startrow=counter, index=True, header=False)
    #     writer.book.save(test_data_path)

    # counter += counter
    # writer = pd.ExcelWriter(self.master_data_sheet, engine='openpyxl')
    # writer.book = book
    # status_dataframe.to_excel(writer, sheet_name='Test_Search_Results')